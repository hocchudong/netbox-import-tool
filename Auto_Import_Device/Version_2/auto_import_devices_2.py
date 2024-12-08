import os
import pandas as pd
import pynetbox
import requests
import re
import urllib3
import warnings
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import load_workbook
import config

filepath = config.filepath
NetBox_URL = config.NetBox_URL
NetBox_Token = config.NetBox_Token
sitename = config.sitename


def file_check(input_file):
    if os.path.exists(filepath):
        workbook = load_workbook(input_file)
        global sheet
        sheet = workbook.active
        data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
        columns = [cell.value for cell in sheet[1]]  # Lấy tiêu đề cột từ dòng đầu tiên
        global df
        df = pd.DataFrame(data, columns=columns)
        df = df.dropna(subset=['Name'], how='all')
        print("File Check complete!")
    else:
        print(f"File '{filepath}' doesn't exist!")
        exit()
     
def netbox_connection_check(netboxurl, netboxtoken):
    try:
        warnings.simplefilter("ignore", InsecureRequestWarning)  
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        response = requests.get(
            netboxurl,
            headers={"Authorization": f"Token {netboxtoken}"},
            timeout=20,
            verify=False  # Tắt xác thực SSL
        )
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        if response.status_code == 200:
            global nb
            nb = pynetbox.api(netboxurl, token=netboxtoken)
            nb.http_session.verify = False  # Tắt xác thực SSL ở đây
            print("Connection Check complete!")
        else:
            print(f"Connection Error: {response.status_code} - {response.reason}")
            return None
    except requests.exceptions.SSLError as e:
        print(f"SSL Error: Can't verify SSL certificate. More: {e}")
    except requests.exceptions.ConnectionError as e:
        print(f"Connection Error: Unable to reach NetBox. More: {e}")
    except requests.exceptions.Timeout as e:
        print(f"Timeout Error: NetBox did not respond in time. More: {e}")
    except requests.exceptions.RequestException as e:
        print(f"Error: An unknown error occurred. More: {e}")
    return None

def excute_merge_data(df):
    try:
        #df = df.dropna(subset=['Name'], how='all')
        df['u_height'] = 1

        # Xử lý các vùng merge
        for merged_cells in sheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_cells.min_row, merged_cells.min_col, merged_cells.max_row, merged_cells.max_col

        # Lấy giá trị từ cột "Name" nếu vùng merge nằm trong cột đó
            if min_col == 3:  # Cột "Name" là cột thứ 3
                merge_height = max_row - min_row + 1
                name_value = sheet.cell(row=min_row, column=min_col).value

                # Lọc ra các dòng có cùng tên với vùng merge
                mask = df['Name'] == name_value

                # Cập nhật cột u_height với chiều cao của vùng merge
                df.loc[mask, 'u_height'] = merge_height

                # Cập nhật Position theo yêu cầu
                if merge_height == 2:
                    df.loc[mask, 'U'] = df.loc[mask, 'U'] - 1
                elif merge_height == 3:
                    df.loc[mask, 'U'] = df.loc[mask, 'U'] - 2
        
        # Chỉ giữ lại một dòng duy nhất cho các dòng merge
        df = df.drop_duplicates(subset=['Name'], keep='first')
        print("Making U_Height complete!")
        return df
    except Exception as e:
        print(f"Error while making u_height: {e}")
        exit()   
           
def device_types_check():
    device_types_in_file = df['Device Type'].dropna().tolist() 
    # Khởi tạo mảng chứa
    device_types_not_in_netbox = []
    # Kiểm tra xem mỗi device type đã có trong NetBox chưa
    for device_type in device_types_in_file:
        search_result = nb.dcim.device_types.filter(model=device_type)
        if not search_result:
            device_types_not_in_netbox.append(device_type)
    device_type_not_in_netbox = []
    for device_type in device_types_not_in_netbox:
        if device_type not in device_type_not_in_netbox:
            device_type_not_in_netbox.append(device_type)
    # In ra danh sách device types chưa có trên NetBox
    if device_type_not_in_netbox:
        print("Device Types not in NetBox:")
        print(device_type_not_in_netbox)

        # Chọn cách xử lý cho các device types chưa có trên NetBox
        print("\nWhat do you want to do with the device_types not in NetBox")
        print("1. Add manual and relaunch later!")
        print("2. Automatic add with sample value")

        choice = input("Enter your choice? (1 or 2): ")

        if choice == '1':
            print("\nPlease Add Device Types manually!")
            exit()
        elif choice == "2":
            print("You chose to Add Device Types Automatically with sample information")
            print("Trying to Add Automatically...")

            # Thêm device types chưa có trong NetBox
            for device_type in device_type_not_in_netbox:
                try:
                    matching_rows = df[df['Device Type'] == device_type]
                    if matching_rows.empty:
                        print(f"Device type {device_type} not found in Excel. Skipping...")
                        continue

                    row = matching_rows.iloc[0]
                    manufacturer_name = row['Manufacturer'].strip()
                    u_height = row['u_height']
                    u_height = int(u_height)
                    # Kiểm tra manufacturer tồn tại hoặc tạo mới
                    # Kiểm tra xem manufacturer đã tồn tại trên NetBox chưa
                    manufacturer_records = nb.dcim.manufacturers.filter(name=manufacturer_name)

                    # Duyệt qua tất cả manufacturers trả về
                    manufacturer = None
                    for record in manufacturer_records:
                        manufacturer = record
                        break  # Lấy manufacturer đầu tiên tìm thấy

                    if manufacturer:
                        print(f"Using existing manufacturer: {manufacturer.name} (ID: {manufacturer.id})")
                    else:
                        # Tạo slug hợp lệ từ manufacturer_name
                        manufacturer_slug = re.sub(r'[^a-z0-9-]', '-', manufacturer_name.lower()).strip('-')

                        # Nếu không có manufacturer, tạo mới
                        manufacturer = nb.dcim.manufacturers.create(
                            name=manufacturer_name,
                            slug=manufacturer_slug  
                        )
                        print(f"Created new manufacturer: {manufacturer.name} (ID: {manufacturer.id})")

                    # Tạo slug hợp lệ cho device type
                    device_type_slug = re.sub(r'[^a-z0-9-]', '-', device_type.lower()).strip('-')

                    # Thêm device type mới
                    new_device_type = nb.dcim.device_types.create({
                        'model': device_type,
                        'slug': device_type_slug,
                        'manufacturer': manufacturer.id,
                        'u_height': u_height,
                        'is_full_depth': 'yes',
                    })
                    print(f"Automatically added: {device_type}")

                except Exception as e:
                    print(f"Error while adding {device_type}: {e}, Data: {row.to_dict()}")
                    exit()
    else:
        print("Device Types check complete!")
        
def site_check(site_name):
    site_record = nb.dcim.sites.get(name=site_name)
    if site_record:
        print("Site check complete!")
    else:
        raise("Site doesn't Exist!")
        
'''
def get_role(role_value):
        if isinstance(role_value, str):  
            if role_value.lower() == 'fw':
                return 'Firewall'
            elif role_value.lower() == 'sw':
                return 'Switch'
            elif role_value.lower() == 'svr':
                return 'Server'
            elif role_value.lower() == 'r':
                return 'Router'
            else:
                return 'Unknown'
        return None  
'''

def device_role_check():
    # Xử lý danh sách device roles từ file Excel
    device_role_names = df['Role'].dropna().tolist()
    
    all_roles_exist = True  # Biến để kiểm tra xem tất cả các role có tồn tại hay không

    for device_roles in device_role_names:
        # Tìm kiếm device role trong NetBox
        dvr = nb.dcim.device_roles.filter(name=device_roles)
        if not dvr:
            print(f"Device Role: {device_roles} does not exist in NetBox, please check again!")
            all_roles_exist = False  # Nếu không tồn tại thì đặt all_roles_exist = False
    
    if all_roles_exist:
        print("Device Role check complete: All roles exist in NetBox!")
    else:
        print("Device Role check complete with missing roles.")
        
def rack_check():
    rack_name = df['Rack'].dropna().drop_duplicates()
    record = nb.dcim.racks.get(name=rack_name)
    if record:
        print("Rack Check complete!")
    else:
        raise("Error while finding rack in NetBox")
    
def get_device_types_ids(device_types_names):
    try:
        device_types = nb.dcim.device_types.filter(name=device_types_names)
        if device_types:
            for device_type in device_types:
                if device_type.model == device_types_names:
                    return device_type.id
        else:
            print(f"Device type '{device_types_names}' not found in NetBox.")
            return None
    except Exception as e:
        print(f"Error fetching device type '{device_types_names}': {e}")
        return None
    
def get_device_roles_ids(device_role_name):
    try:
        device_roles = nb.dcim.device_roles.filter(name=device_role_name)
        if device_roles:
            for device_role in device_roles:
                if device_role.name == device_role_name:
                    return device_role.id
        else:
            print(f"Device Role '{device_role_name}' not found in NetBox.")
            return None
    except Exception as e:
        print(f"Error fetching device role'{device_role_name}': {e}")
        return None
    
def get_site_id(site_name):
    try:
        site = nb.dcim.sites.get(name=site_name)
        return site.id
    except Exception as e:
        print(f"Error fetching site '{site_name}': {e}")
        return None
    
def get_rack_id(rack_name):
    try:
        rack = nb.dcim.racks.get(name=rack_name)
        return rack.id
    except Exception as e:
        print(f"Error fetching rack '{rack_name}': {e}")
        return None
    
def import_device_to_NetBox():
    device_names = df['Name'].dropna().tolist()
    for device_name in device_names:
        matching_rows = df[df['Name'] == device_name]
        if matching_rows.empty:
            print(f"Device Name {device_name} not found in Excel. Skipping...")
            continue
        
        row = matching_rows.iloc[0]
        device_role = row.get('Role')
        name = row.get('Name')
        rack = row.get('Rack')
        device_types = row.get('Device Type')
        serial_number = row.get('Serial Number', ' ')
        comments = row.get('Comments', ' ')
        contract_number = row.get('Contract Number', ' ')
        year_of_investment = row.get('Year of Investment', ' ')
        # Ép kiểu sang dạng int
        position = int(row.get('U'))
        
        device_types_id = get_device_types_ids(device_types)

        device_roles_id = get_device_roles_ids(device_role)

        rack_id = get_rack_id(rack)
        site_id = get_site_id('IDC NTL')
        
        try:
            new_device = nb.dcim.devices.create(
                {
                    "name": name,
                    "device_type":device_types_id,
                    "role": device_roles_id,
                    "site": site_id,
                    "serial": serial_number,
                    "rack": rack_id,
                    "face": "front",
                    "position": position,
                    "status": "active",  
                    "description": comments,
                    "custom_fields":{
                        "contract_number": contract_number,
                        "year_of_investment": year_of_investment
                    }
                }
            )
            print(f"Successfully created device: {name}")
        except Exception as e:
            print(f"Error creating device '{name}': {e}")

def main():
    try:
        # Kiểm tra file đầu vào
        print("Step 1: Checking input file...")
        file_check(filepath)
        
        # Kiểm tra kết nối NetBox
        print("Step 2: Checking NetBox connection...")
        netbox_connection_check(NetBox_URL, NetBox_Token)
        
        # Xử lý data bị merge
        print("Step 3: Excuting Merge data...")
        excute_merge_data(df)
        
        # Kiểm tra Device Types
        print("Step 4: Checking Device Types...")
        device_types_check()

        # Kiểm tra Device Roles
        print("Step 5: Checking Device Roles...")
        device_role_check()

        # Kiểm tra Rack
        print("Step 6: Checking Rack...")
        rack_check()
        
        # Import Devices vào NetBox
        print("Step 7: Importing Devices into NetBox...")
        import_device_to_NetBox()

        print("Process completed successfully!")

    except Exception as e:
        print(f"Error during execution: {e}")

if __name__ == "__main__":
    main()

