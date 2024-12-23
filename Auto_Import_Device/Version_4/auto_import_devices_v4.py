import os
import pandas as pd
import pynetbox
import requests
import re
import urllib3
import warnings
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import load_workbook
from datetime import datetime
import random

WIDTH=19
U_HEIGHT=42
STATUS = 'active'
TAG_NAME_AUTO_IMPORT = "AutoImportExcel"

# FILE_PATH = '/opt/netbox/netbox/plugin/netbox-import-tool/Auto_Import_Device/Version_4/Rack_M1-10.xlsx' 
# FILE_PATH = '/opt/netbox/netbox/plugin/netbox-import-tool/Auto_Import_Device/Version_4/test_rack_nan.xlsx' 
FILE_PATH = '/opt/netbox/netbox/plugin/netbox-import-tool/Auto_Import_Device/Version_4/test_date_time.xlsx'
NetBox_URL = 'http://172.16.66.177:8000'
NetBox_Token = '633a7508b878bcbf33091699289a8a3026a3fbf6'

SITE_NAME = 'VNPT NTL' # Site của  netbox
SHEET_NAME = 'Input' # Tên của sheet muốn import
# Khai báo biến global
TAG_ID_AUTO_IMPORT = []
DEVICE_HEIGHTS =  []
LIST_ADD_DEVICE_ROLE_ERROR = []
LIST_ADD_MANUFACTURES_ERROR = []
LIST_ADD_DEVICE_TYPE_ERROR = []
LIST_ADD_DEVICE_ERROR = []

# Class lưu thông tin độ cao của device type
class DeviceHight:
    def __init__(self, name, height):
        self.name = name
        self.height = height
    
    def __repr__(self):
        return f"DeviceHight(name='{self.name}', height={self.height})"
    
def file_check(input_file):
    if os.path.exists(input_file):
        # read file excel
        workbook = load_workbook(input_file)
        global sheet
        sheet = workbook.active
        global df
        columns = [cell.value for cell in sheet[1]]  
        # new code
        df = pd.read_excel(input_file, sheet_name=SHEET_NAME)
    else:
        print(f"File '{input_file}' doesn't exist!")
        exit()

def netbox_connection_check(netboxurl, netboxtoken):
    try:
        warnings.simplefilter("ignore", InsecureRequestWarning)  
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        response = requests.get(
            netboxurl,
            headers={"Authorization": f"Token {netboxtoken}"},
            timeout=20,
            verify=False  
        )
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        if response.status_code == 200:
            global nb
            nb = pynetbox.api(netboxurl, token=netboxtoken)
            nb.http_session.verify = False  
            print("Connection Check complete!")
        else:
            print(f"Connection Error: {response.status_code} - {response.reason}")
            return None
    except requests.exceptions.SSLError as e:
        print(f"SSL Error: Can't verify SSL certificate. More: {e}")
    except requests.exceptions.ConnectionError as e:
        print(f"Connection Error: Unable to reach NetBox. More: {e}")
    except requests.exceptions.Timeout as e:
        print(f"Timeout Error: NetBox did not respond in time. More: {e}")
    except requests.exceptions.RequestException as e:
        print(f"Error: An unknown error occurred. More: {e}")
    return None

def site_check(site_name):
    site_record = nb.dcim.sites.get(name=site_name)
    if site_record:
        print("Site check complete!")
    else:
        print(f"Site '{site_name}' does not exist in NetBox!") 
        print("Do you want to auto add new Site?")
        choice = input("yes/no: ").strip().lower()
        if choice == 'yes':
            new_site = nb.dcim.sites.create(
                name=site_name,
                slug=site_name.lower().replace(" ", "-"),
                tags=TAG_ID_AUTO_IMPORT,
                status=STATUS,
                description='Create by Auto_Import_Tool',
            )
            print(f"Successfully created Site: {site_name}")
        else:
            print("Please check the Site again before using this Tool!")


# Hàm check xem có tag AutoImportExcel chưa
def tag_check():
    # get tag 
    tag_exist = nb.extras.tags.filter(slug="auto-import-excel")

    # if not exist => add tag AutoImportExcel
    if not tag_exist:
        try:
            new_tag = nb.extras.tags.create(
                name=TAG_NAME_AUTO_IMPORT,
                slug="auto-import-excel",
                color="9e9e9e",
                description='Create tag AutoImportExcel',
            )
            print(f"Create success tag {TAG_NAME_AUTO_IMPORT}")
            # get ID tag 
            get_tag_id()
        except:
            print(f"Error while create tag {TAG_NAME_AUTO_IMPORT}")
    else:
        # get ID tag
        get_tag_id()

def get_tag_id():
    tag_auto_import_excel = nb.extras.tags.filter(name=TAG_NAME_AUTO_IMPORT)
    first_tag = next(tag_auto_import_excel, None)
    if first_tag:
        TAG_ID_AUTO_IMPORT.append(first_tag.id)

def get_role(role_value):
        if isinstance(role_value, str):  
            if role_value.lower() == 'fw':
                return 'Firewall'
            elif role_value.lower() == 'sw':
                return 'Switch'
            elif role_value.lower() == 'svr':
                return 'Server'
            elif role_value.lower() == 'r':
                return 'Router'
            else:
                return 'Unknown'
        return None  

def device_role_check():
    # device_role_names = df['Role'].dropna().drop_duplicates().apply(get_role).tolist()
    device_role_names = df["Role"].tolist() # chuyển pandas thành list 
    device_role_names = list(set(device_role_names)) # lọc trùng
    device_role_names = list(filter(lambda x: x == x, device_role_names)) # loại bỏ phần tử nan trong chuỗi

    all_roles_exist = True # tất cả đã tồn tại chưa
    missing_roles = [] # danh sách các roles chưa có 

    # tìm kiếm các role bị thiếu
    for device_roles in device_role_names:
        dvr = nb.dcim.device_roles.filter(name=device_roles)
        if not dvr:
            print(f"Device Role: {device_roles} does not exist in NetBox, please check again!")
            all_roles_exist = False  
            missing_roles.append(device_roles)
    
    # kiểm tra tất cả role đã tồn tại hay chưa
    if all_roles_exist:
        print("Device Role check complete: All roles exist in NetBox!")
    else:
        print("Do you want to auto add Role?")
        choice = input("yes/no: ").strip().lower()

        if choice == 'yes':
            for role in missing_roles:
                try:
                    new_role = nb.dcim.device_roles.create(
                        name=role,
                        slug=role.lower().replace(" ", "-"),
                        tags=TAG_ID_AUTO_IMPORT,
                        color="9e9e9e",
                        description='Create device role by Auto_Import_Tool',  
                    )
                    print(f"Successfully created Device Role: {new_role['name']}")
                except Exception as e:
                    LIST_ADD_DEVICE_ROLE_ERROR.append(role)
                    print(f"Failed to create Device Role {role}: {e}")
        else:
            print("No roles were added. Please update NetBox manually if needed.")
            exit()
        
def rack_check():
    rack_names = df['Rack'].drop_duplicates().dropna().tolist()
    rack_names = list(set(rack_names))# lọc trùng
    missing_racks = []
    for rack_name in rack_names:
        record = nb.dcim.racks.get(name=rack_name)
        if record:
            continue
        else:
            print(f"Rack '{rack_name}' does not exist in NetBox!")
            missing_racks.append(rack_name)
    
    if missing_racks:
        print("Do you want to auto add new Rack?")
        choice = input("yes/no: ").strip().lower()
        if choice == 'yes':
            for rack in missing_racks:
                try:
                    site = nb.dcim.sites.get(name=SITE_NAME)
                    if not site:
                        print(f"Site '{SITE_NAME}' does not exist in NetBox. Please create it first.")
                        break
                    new_rack = nb.dcim.racks.create(
                        site=site.id,
                        name=rack,
                        status='active',
                        tags=TAG_ID_AUTO_IMPORT,
                        width=WIDTH,
                        u_height=U_HEIGHT,
                        description='Create by Auto_Import_Tool',
                    )
                    print(f"Successfully created Rack: {new_rack['name']}")
                except Exception as e:
                    print(f"Failed to create Rack {rack}: {e}")
        else:
            print("No racks were added. Please update NetBox manually if needed.")
    else:
        print("Rack Check complete: All racks exist in NetBox!")

def manufacturer_check():
    manufacturer_names = df["Manufacturer"].dropna().tolist()
    manufacturer_names = list(set(manufacturer_names))# lọc trùng

    all_manufacturers_exist = True  
    missing_manufacturers = []
    for manufacturer in manufacturer_names:
        dvr = nb.dcim.manufacturers.filter(name=manufacturer)
        if not dvr:
            print(f"Manufacturer: {manufacturer} does not exist in NetBox, please check again!")
            all_manufacturers_exist = False  
            missing_manufacturers.append(manufacturer)

    if all_manufacturers_exist:
        print("Device manufacturer check complete: All manufacturers exist in NetBox!")
    else:
        print("Do you want to auto add manufacturer?")
        choice = input("yes/no: ").strip().lower()

        if choice == 'yes':
            for manufacturer in missing_manufacturers:
                try:
                    new_manufacturer = nb.dcim.manufacturers.create(
                        name=manufacturer,
                        slug=manufacturer.lower().replace(" ", "-"),
                        tags=TAG_ID_AUTO_IMPORT,
                        description='Create manufacture by Auto_Import_Tool',  
                    )
                    print(f"Successfully created Manufacture: {new_manufacturer['name']}")
                except Exception as e:
                    LIST_ADD_MANUFACTURES_ERROR.append(manufacturer)
                    print(f"Failed to create Manufacture {manufacturer}: {e}")
        else:
            print("No manufacturers were added. Please update NetBox manually if needed.")
            exit()

def custom_feild_check():
    custom_feild_names = ["device_owner", "contract_number" ,"year_of_investment"]
    all_custom_feild_exist = True  
    missing_custom_feild = []
    for custom_feild in custom_feild_names:
        dvr = nb.extras.custom_fields.filter(name=custom_feild)
        if not dvr:
            print(f"custom_feild: {custom_feild} does not exist in NetBox, please check again!")
            all_custom_feild_exist = False  
            missing_custom_feild.append(custom_feild)

    if all_custom_feild_exist:
        print("Device custom_feild check complete: All custom_feild exist in NetBox!")
    else:
        print("Do you want to auto add custom_feild?")
        choice = input("yes/no: ").strip().lower()

        if choice == 'yes':
            for custom_feild in missing_custom_feild:
                try:
                    
                    type_custom_feild = "text"
                    new_custom_feild = nb.extras.custom_fields.create({
                        "name" : custom_feild,
                        "type" : type_custom_feild,
                        "object_types" : ["dcim.device"],
                        "search_weight" : 1000,
                        "weight" : 100,
                        "tags": TAG_ID_AUTO_IMPORT,
                        "filter_logic" : "loose",
                        "ui_visible" : "always",
                        "ui_editable" : "yes",
                        "description" : 'Create by Auto_Import_Tool',
                    })
                    print(f"Successfully created custom feild : {new_custom_feild['name']}")
                except Exception as e:
                    print(f"Failed to create custom feild {custom_feild}: {e}")
        else:
            print("No custom_feild were added. Please update NetBox manually if needed.")
            exit()

#Hàm tạo ra mảng chứa trường name và u_height của device_type
def device_type_height():
    # Tìm u_height của device type 
    print("Merged cell ranges in column H spanning rows:")
    for merged_range in sheet.merged_cells.ranges:
        # Lấy giá trị ở cột  H
        if merged_range.min_col == 8 and merged_range.max_col != merged_range.max_row:
            # lấy value của ô 
            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            device_type_name = top_left_cell.value

            # Tính các row đã merg của ô
            rows_spanned = merged_range.max_row - merged_range.min_row + 1

            exists = any(device.name == device_type_name for device in DEVICE_HEIGHTS)
            # kiểm tra giá trị đã có trong list chưa
            if not exists:
                new_device_height =  DeviceHight(device_type_name,rows_spanned)
                DEVICE_HEIGHTS.append(new_device_height)     

# Hàm auto add device type vào netbox
def device_types_check():
    device_types_in_file = df['Type'].dropna().tolist() 
    device_types_in_file = list(set(device_types_in_file)) # lọc trùng
    device_type_not_in_netbox = []

    # Tìm phần tử chưa có trong netbox
    for device_type in device_types_in_file:
        search_result = nb.dcim.device_types.filter(model=device_type)
        if not search_result:
            device_type_not_in_netbox.append(device_type)

    if device_type_not_in_netbox:
        print("Device Types not in NetBox:")
        print(device_type_not_in_netbox)
        print("\nDo you want to add Device Type automatically?")
        choice = input("Enter your choice? (yes/no): ")

        if choice == 'no':
            print("\nPlease Add Device Types manually!")
            exit()
        elif choice == "yes":
            print("You chose to Add Device Types Automatically with sample information")
            print("Trying to Add Automatically...")
            for device_type in device_type_not_in_netbox:
                try:
                    # Tìm manufature của device type trong danh sách 
                    all_manufacturer_device_type = df[["Manufacturer","Type"]].dropna()
                    manufacturer_device_type = all_manufacturer_device_type.query(f"Type == '{device_type}'")
                    manufacturer_name = manufacturer_device_type.iloc[0]['Manufacturer']
                    manufacturer_name = manufacturer_name.strip()
                    manufacturer = nb.dcim.manufacturers.get(name = manufacturer_name)

                    # Kiểm tra xem có tìm thấy manuyfacturers không
                    if not manufacturer: 
                        print(f"Error while adding {device_type}: Not found manufactures")
                        continue
                    # Lấy chiều cao trong list đã tạo
                    device_height = 1 # mặc định height bằng 1 
                    
                    #Tìm kiếm trong danh sách thiết bị có height > 1
                    for device in DEVICE_HEIGHTS:
                        if device.name == device_type:
                            device_height = device.height
                            break
                    
                    # create device type
                    device_type_slug = re.sub(r'[^a-z0-9-]', '-', device_type.lower()).strip('-')
                    new_device_type = nb.dcim.device_types.create({
                        'model': device_type,
                        'slug': device_type_slug,
                        'manufacturer': manufacturer.id,
                        "tags": TAG_ID_AUTO_IMPORT,
                        'u_height': device_height,
                        'is_full_depth': 'yes',
                    })

                    print(f"Automatically added : {new_device_type}")
                    
                except Exception as e:
                    print(f"Error while adding {device_type}: {e}")
                    LIST_ADD_DEVICE_TYPE_ERROR.append(device_type)
    else:
        print("Device Types check complete!")
    
def get_device_types_ids(device_types_names):
    try:
        device_types_names = device_types_names.strip()
        device_types = nb.dcim.device_types.filter(name=device_types_names)
        if device_types:
            for device_type in device_types:
                if device_type.model == device_types_names:
                    return device_type.id
        else:
            print(f"Device type '{device_types_names}' not found in NetBox.")
            return None
    except Exception as e:
        print(f"Error fetching device type '{device_types_names}': {e}")
        return None
    
def get_device_roles_ids(device_role_name):
    try:
        device_role_name = device_role_name.strip()
        device_roles = nb.dcim.device_roles.filter(name=device_role_name)
        if device_roles:
            for device_role in device_roles:
                if device_role.name == device_role_name:
                    return device_role.id
        else:
            print(f"Device Role '{device_role_name}' not found in NetBox.")
            return None
    except Exception as e:
        print(f"Error fetching device role'{device_role_name}': {e}")
        return None
    
def get_site_id(site_name):
    try:
        site = nb.dcim.sites.get(name=site_name)
        return site.id
    except Exception as e:
        print(f"Error fetching site '{site_name}': {e}")
        return None
    
def get_rack_id(rack_name):
    try:
        rack_name = rack_name.strip()
        rack = nb.dcim.racks.get(name=rack_name)
        return rack.id
    except Exception as e:
        print(f"Error fetching rack '{rack_name}': {e}")
        return None

# Hàm auto add device vào netbox
def import_device_to_NetBox():
    # list device need add 
    device_names = df[["Rack", "U", "Manufacturer", "Name", "Role", "Owner Device",  "Contract number","Type", "Serial Number","Year of Investment", "Description"]]
    device_names= device_names[device_names['Name'].notna()] # lọc tất cả hàng có trường Name là nan
    number_of_device_in_file = 0 # Số lượng device đã được add 
    number_of_device_has_been_added = 0 # Tổng số lượn device trong danh sách

    # check device have serial number null
    device_null_serial_numbers = device_names[device_names['Serial Number'].isnull()]
    if len(device_null_serial_numbers) > 0:
        print(f"List device have serial_number null:\n {device_null_serial_numbers}")
        print("Do you want add serial_number null?")

        choice = input("Enter your choice? (yes/no): ")
        if choice == "yes":
            print("\nPlease Add Device Types manually!")
            exit()

    for index, row in device_names.iterrows():
        number_of_device_in_file+=1

        try:

            # Thêm mới device đến netbox
            rack_id = get_rack_id(row['Rack'])
            site_id = get_site_id(site_name=SITE_NAME)
            device_types_id = get_device_types_ids(row['Type'])
            device_roles_id = get_device_roles_ids(row['Role'])

            # Xử lý trường position
            device_position = row['U']
            #Tìm kiếm trong danh sách thiết bị có height > 1
            for device in DEVICE_HEIGHTS:
                if device.name == row['Type']:
                    device_position = device_position - device.height + 1
                    break

            #Kiểm tra nan cho các param 
            device_description = ""
            device_owner = ""
            contract_number = ""
            device_year_of_investment = ""
            device_serial_number = ""

            if not pd.isna(row['Description']):
                device_description = row['Description']

            if not pd.isna(row['Owner Device']):
                device_owner = row['Owner Device']

            if not pd.isna(row['Contract number']):
                contract_number = row['Contract number']

            if not pd.isna(row['Year of Investment']):
                device_year_of_investment = row['Year of Investment']
                if type(device_year_of_investment) is datetime:
                    device_year_of_investment = device_year_of_investment.strftime("%d-%m-%Y")
                elif type(device_year_of_investment) is int:
                    device_year_of_investment = str(device_year_of_investment)

            # Nếu như không phải null thì sẽ lấy giá trị đó (Mặc định là random)
            if not pd.isna(row['Serial Number']):
                device_serial_number = row['Serial Number']
            else:
                random_number = random.randint(100000, 999999)
                device_serial_number = random_number

            # kiểm tra xem device name đã được add trên netbox chưa
            device_name = row['Name']
            device_name = device_name.strip()
            exist_device = nb.dcim.devices.get(name=device_name)
            if exist_device:
                device_name = device_name + f"-{device_serial_number}"

            new_device = nb.dcim.devices.create(
                {
                    "name": device_name,
                    "device_type": device_types_id,
                    "role": device_roles_id,
                    "site": site_id,
                    "serial": device_serial_number,
                    "rack": rack_id,
                    "face": "front",
                    "position": device_position,
                    "tags":TAG_ID_AUTO_IMPORT,
                    "status": STATUS,  
                    "description": device_description,
                    "custom_fields": {
                        "device_owner": device_owner,
                        "contract_number": contract_number,
                        "year_of_investment": device_year_of_investment,
                    }
                }
            )
            
            number_of_device_has_been_added+=1
            print(f"Successfully created device: {device_name}")
        except Exception as e:
            LIST_ADD_DEVICE_ERROR.append(device_name)
            print(f"Error creating device '{device_name}': {e}")
            
    if number_of_device_has_been_added > 0:
        print(f"{number_of_device_has_been_added}/{number_of_device_in_file} device has been added to NetBox!" )

def main():
    try:
        print("Step 1: Checking input file...")
        file_check(FILE_PATH)
        
        print("Step 2: Checking NetBox connection...")
        netbox_connection_check(NetBox_URL, NetBox_Token)

        print(f"Step 3: Check tag {TAG_NAME_AUTO_IMPORT}")
        tag_check()

        print("Step 4: check site name")
        site_check(site_name=SITE_NAME)

        print("Step 5: check rack")
        rack_check()

        print("Step 6: check device role")
        device_role_check()
        
        print("Step 7: check manufacturer")
        manufacturer_check()

        print("Step 8: check custom feild")
        custom_feild_check()
        
        print("Step 9: Check height device type")
        device_type_height()

        print("Step 10: check device type")
        device_types_check()

        print("Step 11: Importing Devices into NetBox...")
        import_device_to_NetBox()

        print(f"List Manufacture error while create new record:\n {LIST_ADD_MANUFACTURES_ERROR}")

        print(f"List Device Type error while create new record:\n {LIST_ADD_DEVICE_TYPE_ERROR}")

        print(f"List Device Role error while create new record:\n {LIST_ADD_DEVICE_ROLE_ERROR}")

        print(f"List Device error while create new record:\n {LIST_ADD_DEVICE_ERROR}")

        print("Process completed successfully!")
    except Exception as e:
        print(f"Error during execution: {e}")

if __name__ == "__main__":
    main()

