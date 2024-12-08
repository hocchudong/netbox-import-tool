import os
import pandas as pd
import pynetbox
import requests
import re
import urllib3
import warnings
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import load_workbook
import config

filepath = config.filepath
NetBox_URL = config.NetBox_URL
NetBox_Token = config.NetBox_Token
sitename = config.sitename

def file_check(input_file):
    if os.path.exists(input_file):
        workbook = load_workbook(input_file)
        global sheet
        sheet = workbook.active
        data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
        columns = [cell.value for cell in sheet[1]]  
        global df
        df = pd.DataFrame(data, columns=columns)
        df = df.dropna(subset=['Name'], how='all')
        required_columns = ['Role','Type','Serial Number']
        for index, row in df.iterrows():
            if pd.notna(row['Name']):
                missing_columns = [col for col in required_columns if pd.isnull(row[col])]
                if missing_columns:
                    print(f"Error: Row {index + 2} is missing values in columns: {missing_columns}")
                    exit()
        print("File Check complete!")
    else:
        print(f"File '{input_file}' doesn't exist!")
        exit()
    
def netbox_connection_check(netboxurl, netboxtoken):
    try:
        warnings.simplefilter("ignore", InsecureRequestWarning)  
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        response = requests.get(
            netboxurl,
            headers={"Authorization": f"Token {netboxtoken}"},
            timeout=20,
            verify=False  
        )
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        if response.status_code == 200:
            global nb
            nb = pynetbox.api(netboxurl, token=netboxtoken)
            nb.http_session.verify = False  
            print("Connection Check complete!")
        else:
            print(f"Connection Error: {response.status_code} - {response.reason}")
            return None
    except requests.exceptions.SSLError as e:
        print(f"SSL Error: Can't verify SSL certificate. More: {e}")
    except requests.exceptions.ConnectionError as e:
        print(f"Connection Error: Unable to reach NetBox. More: {e}")
    except requests.exceptions.Timeout as e:
        print(f"Timeout Error: NetBox did not respond in time. More: {e}")
    except requests.exceptions.RequestException as e:
        print(f"Error: An unknown error occurred. More: {e}")
    return None

def handle_duplicate_names(df, name_col, serial_col):
    df.columns = df.columns.str.strip()
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
    name_counts = df[name_col].value_counts()
    duplicates = name_counts[name_counts > 1].index  
    for name in duplicates:
        duplicate_rows = df[df[name_col] == name]
        for row in duplicate_rows.index:
            serial_value = df.at[row,serial_col]
            df.at[row, name_col] = f"{name}_{serial_value}"
    return df  

def excute_merge_data(df):
    try:
        df['u_height'] = 1
        for merged_cells in sheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_cells.min_row, merged_cells.min_col, merged_cells.max_row, merged_cells.max_col
            if min_col == 3:  
                merge_height = max_row - min_row + 1
                name_value = sheet.cell(row=min_row, column=min_col).value
                mask = df['Name'] == name_value
                df.loc[mask, 'u_height'] = merge_height
                if merge_height > 1:
                    df.loc[mask, 'Position'] = df.loc[mask, 'Position'] - (merge_height-1)
        df = df.drop_duplicates(subset=['Name'], keep='first')
        print("Merge data has been excuted!")
        return df
    except Exception as e:
        print(f"Error while excuting merge_data: {e}")
        exit()   
           
def device_types_check():
    device_types_in_file = df['Type'].dropna().tolist() 
    device_types_not_in_netbox = []
    for device_type in device_types_in_file:
        search_result = nb.dcim.device_types.filter(model=device_type)
        if not search_result:
            device_types_not_in_netbox.append(device_type)
    device_type_not_in_netbox = []
    for device_type in device_types_not_in_netbox:
        if device_type not in device_type_not_in_netbox:
            device_type_not_in_netbox.append(device_type)
    if device_type_not_in_netbox:
        print("Device Types not in NetBox:")
        print(device_type_not_in_netbox)
        print("\nDo you want to add Device Type automatically?")
        print("1. Add manual and relaunch later!")
        print("2. Automatic add with sample value")
        choice = input("Enter your choice? (1 or 2): ")
        if choice == '1':
            print("\nPlease Add Device Types manually!")
            exit()
        elif choice == "2":
            print("You chose to Add Device Types Automatically with sample information")
            print("Trying to Add Automatically...")
            for device_type in device_type_not_in_netbox:
                try:
                    matching_rows = df[df['Type'] == device_type]
                    if matching_rows.empty:
                        print(f"Device type {device_type} not found in Excel. Skipping...")
                        continue
                    row = matching_rows.iloc[0]
                    manufacturer_name = row['Manufacturer'].strip()
                    u_height = row['u_height']
                    u_height = int(u_height)
                    manufacturer_records = nb.dcim.manufacturers.filter(name=manufacturer_name)
                    manufacturer = None
                    for record in manufacturer_records:
                        manufacturer = record
                        break  
                    if manufacturer:
                        print(f"Using existing manufacturer: {manufacturer.name} (ID: {manufacturer.id})")
                    else:
                        manufacturer_slug = re.sub(r'[^a-z0-9-]', '-', manufacturer_name.lower()).strip('-')
                        manufacturer = nb.dcim.manufacturers.create(
                            name=manufacturer_name,
                            slug=manufacturer_slug  
                        )
                        print(f"Created new manufacturer: {manufacturer.name} (ID: {manufacturer.id})")
                    device_type_slug = re.sub(r'[^a-z0-9-]', '-', device_type.lower()).strip('-')
                    new_device_type = nb.dcim.device_types.create({
                        'model': device_type,
                        'slug': device_type_slug,
                        'manufacturer': manufacturer.id,
                        'u_height': u_height,
                        'is_full_depth': 'yes',
                    })
                    print(f"Automatically added: {device_type}")
                except Exception as e:
                    print(f"Error while adding {device_type}: {e}, Data: {row.to_dict()}")
                    exit()
    else:
        print("Device Types check complete!")
        
def site_check(site_name):
    site_record = nb.dcim.sites.get(name=site_name)
    if site_record:
        print("Site check complete!")
    else:
        print(f"Site '{site_name}' does not exist in NetBox!") 
        print("Do you want to auto add new Site?")
        choice = input("yes/no: ").strip().lower()
        if choice == 'yes':
            new_site = nb.dcim.sites.create(
                name=site_name,
                slug=site_name.lower().replace(" ", "-"),
                status=config.status,
                description='Create by Auto_Import_Tool',
            )
            print(f"Successfully created Site: {site_name}")
        else:
            print("Please check the Site again before using this Tool!")
            
def get_role(role_value):
        if isinstance(role_value, str):  
            if role_value.lower() == 'fw':
                return 'Firewall'
            elif role_value.lower() == 'sw':
                return 'Switch'
            elif role_value.lower() == 'svr':
                return 'Server'
            elif role_value.lower() == 'r':
                return 'Router'
            else:
                return 'Unknown'
        return None  

def device_role_check():
    device_role_names = df['Role'].dropna().drop_duplicates().apply(get_role).tolist()
    all_roles_exist = True  
    missing_roles = []
    for device_roles in device_role_names:
        dvr = nb.dcim.device_roles.filter(name=device_roles)
        if not dvr:
            print(f"Device Role: {device_roles} does not exist in NetBox, please check again!")
            all_roles_exist = False  
            missing_roles.append(device_roles)
    if all_roles_exist:
        print("Device Role check complete: All roles exist in NetBox!")
    else:
        print("Do you want to auto add Role?")
        choice = input("yes/no: ").strip().lower()

        if choice == 'yes':
            for role in missing_roles:
                try:
                    new_role = nb.dcim.device_roles.create(
                        name=role,
                        slug=role.lower().replace(" ", "-"),
                        color="9e9e9e",
                        description='Create by Auto_Import_Tool',  
                    )
                    print(f"Successfully created Device Role: {new_role['name']}")
                except Exception as e:
                    print(f"Failed to create Device Role {role}: {e}")
        else:
            print("No roles were added. Please update NetBox manually if needed.")
            exit()
        
def rack_check():
    rack_names = df['Rack'].drop_duplicates().tolist()
    missing_racks = []
    for rack_name in rack_names:
        record = nb.dcim.racks.get(name=rack_name)
        if record:
            continue
        else:
            print(f"Rack '{rack_name}' does not exist in NetBox!")
            missing_racks.append(rack_name)
    if missing_racks:
        print("Do you want to auto add new Rack?")
        choice = input("yes/no: ").strip().lower()
        if choice == 'yes':
            for rack in missing_racks:
                try:
                    site = nb.dcim.sites.get(name=config.sitename)
                    if not site:
                        print(f"Site '{config.sitename}' does not exist in NetBox. Please create it first.")
                        break
                    new_rack = nb.dcim.racks.create(
                        site=site.id,
                        name=rack,
                        status='active',
                        width=config.width,
                        u_height=config.u_height,
                        description='Create by Auto_Import_Tool',
                    )
                    print(f"Successfully created Rack: {new_rack['name']}")
                except Exception as e:
                    print(f"Failed to create Rack {rack}: {e}")
        else:
            print("No racks were added. Please update NetBox manually if needed.")
    else:
        print("Rack Check complete: All racks exist in NetBox!")
   
def get_device_types_ids(device_types_names):
    try:
        device_types = nb.dcim.device_types.filter(name=device_types_names)
        if device_types:
            for device_type in device_types:
                if device_type.model == device_types_names:
                    return device_type.id
        else:
            print(f"Device type '{device_types_names}' not found in NetBox.")
            return None
    except Exception as e:
        print(f"Error fetching device type '{device_types_names}': {e}")
        return None
    
def get_device_roles_ids(device_role_name):
    try:
        device_roles = nb.dcim.device_roles.filter(name=device_role_name)
        if device_roles:
            for device_role in device_roles:
                if device_role.name == device_role_name:
                    return device_role.id
        else:
            print(f"Device Role '{device_role_name}' not found in NetBox.")
            return None
    except Exception as e:
        print(f"Error fetching device role'{device_role_name}': {e}")
        return None
    
def get_site_id(site_name):
    try:
        site = nb.dcim.sites.get(name=site_name)
        return site.id
    except Exception as e:
        print(f"Error fetching site '{site_name}': {e}")
        return None
    
def get_rack_id(rack_name):
    try:
        rack = nb.dcim.racks.get(name=rack_name)
        return rack.id
    except Exception as e:
        print(f"Error fetching rack '{rack_name}': {e}")
        return None
    
def import_device_to_NetBox(df):
    df = handle_duplicate_names(df,name_col='Name', serial_col='Serial Number')
    device_names = df['Name'].dropna().tolist()
    number_of_device_in_file = 0
    number_of_device_has_been_added = 0
    df['role'] = df['Role'].apply(get_role)
    df['description'] = df['Description'].fillna('No data yet')
    for device_name in device_names:
        matching_rows = df[df['Name'] == device_name]
        if matching_rows.empty:
            print(f"Device Name {device_name} not found in Excel. Skipping...")
            continue
        number_of_device_in_file += 1
        row = matching_rows.iloc[0]
        device_role = row.get('role')
        name = row.get('Name')
        rack = row.get('Rack')
        device_types = row.get('Type')
        serial_number = row.get('Serial Number', ' ')
        description = row.get('description', ' ')
        contract_number = row.get('Contract Number', ' ')
        year_of_investment = row.get('Year of Investment', ' ')
        device_owner = row.get('Device Owner', ' ')
        position = int(row.get('Position'))
        
        device_types_id = get_device_types_ids(device_types)
        device_roles_id = get_device_roles_ids(device_role)
        rack_id = get_rack_id(rack)
        site_id = get_site_id(site_name=sitename)
        
        try:
            new_device = nb.dcim.devices.create(
                {
                    "name": name,
                    "device_type": device_types_id,
                    "role": device_roles_id,
                    "site": site_id,
                    "serial": serial_number,
                    "rack": rack_id,
                    "face": "front",
                    "position": position,
                    "status": config.status,  
                    "description": description,
                    "custom_fields": {
                        "device_owner": device_owner,
                        "contract_number": contract_number,
                        "years_of_investment": year_of_investment
                    }
                }
            )
            number_of_device_has_been_added+=1
            print(f"Successfully created device: {name}")
        except Exception as e:
            print(f"Error creating device '{name}': {e}")
    if number_of_device_has_been_added > 0:
        print(f"{number_of_device_has_been_added}/{number_of_device_in_file} device has been added to NetBox!" )

def main():
    try:
        print("Step 1: Checking input file...")
        file_check(filepath)
        
        print("Step 2: Checking NetBox connection...")
        netbox_connection_check(NetBox_URL, NetBox_Token)
        
        print("Step 3: Excuting Merge data...")
        excute_merge_data(df)

        print("Step 4: Checking data in NetBox...")
        site_check(site_name=sitename)
        rack_check()
        device_role_check()
        device_types_check()

        print("Step 5: Importing Devices into NetBox...")
        import_device_to_NetBox(df)

        print("Process completed successfully!")

    except Exception as e:
        print(f"Error during execution: {e}")

if __name__ == "__main__":
    main()

