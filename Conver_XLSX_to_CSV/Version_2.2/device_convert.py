import pandas as pd
from openpyxl import load_workbook
import config

file_path = config.file_path
workbook = load_workbook(file_path)
sheet = workbook.active
data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]
columns = [cell.value for cell in sheet[1]]  
df = pd.DataFrame(data, columns=columns)
df = df.dropna(subset=['Name'], how='all')

for merged_cells in sheet.merged_cells.ranges:
    min_row, min_col, max_row, max_col = merged_cells.min_row, merged_cells.min_col, merged_cells.max_row, merged_cells.max_col
    if min_col == 3:  
        merge_height = max_row - min_row + 1
        name_value = sheet.cell(row=min_row, column=min_col).value
        mask = df['Name'] == name_value
        if merge_height == 2:
            df.loc[mask, 'U'] = df.loc[mask, 'U'] - 1
        elif merge_height == 3:
            df.loc[mask, 'U'] = df.loc[mask, 'U'] - 2
df = df.drop_duplicates(subset=['Name'], keep='first')

def get_role(role_value):
    if isinstance(role_value, str):  
        if role_value.lower() == 'fw':
            return 'Firewall'
        elif role_value.lower() == 'sw':
            return 'Switch'
        elif role_value.lower() == 'svr':
            return 'Server'
        elif role_value.lower() == 'router':
            return 'Router'
        else:
            return role_value
    return None  

def handle_duplicate_names(df, name_col, rack_col, position_col):
    name_counts = df[name_col].value_counts()
    duplicates = name_counts[name_counts > 1].index  
    for name in duplicates:
        duplicate_rows = df[df[name_col] == name]
        for row in duplicate_rows.index:
            rack_value = df.at[row, rack_col]
            position_value = df.at[row, position_col]
            df.at[row, name_col] = f"{name}_{rack_value}_U{position_value}"
    return df

df['role'] = df['Role'].apply(get_role)
df = df.dropna(subset=['Role', 'role'])
df = handle_duplicate_names(df, name_col='Name', rack_col='Rack', position_col='U')
output_columns = [
    'role', 'manufacturer', 'device_type', 'status', 'site', 'name',
    'serial', 'rack', 'position', 'face', 'comments',
    'cf_contract_number', 'cf_years_of_investment',
]

df_csv = pd.DataFrame(columns=output_columns)
df_csv['role'] = df['role']
df_csv['manufacturer'] = df['Manufacturer']
df_csv['device_type'] = df['Device Type']
df_csv['serial'] = df['Serial Number']
df_csv['name'] = df['Name']
df_csv['position'] = df['U']
df_csv['cf_years_of_investment'] = df['Year of Investment']
df_csv['comments'] = df['Comments']
df_csv['cf_contract_number'] = df['Contract Number']
df_csv['rack'] = df['Rack']
df_csv['status'] = config.status      
df_csv['site'] = config.site     
df_csv['face'] = config.face    

     
df_csv.to_csv(config.output_file_path, index=False)
print(f"Saving file.csv successfully at: {config.output_file_path}")
